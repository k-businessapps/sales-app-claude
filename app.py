import base64
import json
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# Branding
# =========================
EMAIL_REGEX = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)

KC_PRIMARY = "#B04EF0"
KC_ACCENT = "#E060F0"
KC_DEEP = "#8030F0"
KC_SOFT = "#F6F0FF"

EXCLUDED_OWNER_CANON = "pipedrive krispcall"
CREDIT_EXCLUDE_DESCS = {"purchased credit", "credit purchased", "amount recharged"}


# =========================
# Secrets + UI helpers
# =========================
def _get_secret(path: List[str], default=None):
    cur = st.secrets
    for key in path:
        if key not in cur:
            return default
        cur = cur[key]
    return cur


def _inject_brand_css():
    st.markdown(
        f"""
        <style>
          .kc-hero {{ padding: 18px 18px; border-radius: 18px; background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PRIMARY} 45%, {KC_ACCENT} 100%); color: white; box-shadow: 0 10px 30px rgba(0,0,0,0.08); }}
          .kc-hero h1 {{ margin: 0; font-size: 28px; line-height: 1.2; }}
          .kc-hero p {{ margin: 6px 0 0 0; opacity: 0.95; font-size: 14px; }}
          .kc-card {{ background: white; border: 1px solid rgba(176, 78, 240, 0.18); border-radius: 16px; padding: 14px 14px; box-shadow: 0 10px 24px rgba(20, 6, 31, 0.04); }}
          .kc-muted {{ color: rgba(20, 6, 31, 0.72); }}
          div.stButton > button {{ border-radius: 14px !important; border: 0 !important; background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PRIMARY} 55%, {KC_ACCENT} 100%) !important; color: white !important; padding: 0.55rem 1rem !important; font-weight: 600 !important; }}
          section[data-testid="stFileUploaderDropzone"] {{ border-radius: 14px; border: 2px dashed rgba(176, 78, 240, 0.35); background: {KC_SOFT}; }}
          div[data-testid="stDataFrame"] {{ border-radius: 14px; overflow: hidden; }}
          .block-container {{ padding-top: 1.1rem; padding-bottom: 1.2rem; }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _logo_html(width_px: int = 220, top_pad_px: int = 10) -> str:
    logo_path = Path(__file__).parent / "assets" / "KrispCallLogo.png"
    if not logo_path.exists():
        return ""
    b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
    return f'<div style="padding-top:{top_pad_px}px;"><img src="data:image/png;base64,{b64}" style="width:{width_px}px; height:auto;" /></div>'


def require_login():
    st.session_state.setdefault("authenticated", False)
    if st.session_state["authenticated"]:
        return

    _inject_brand_css()
    c1, c2 = st.columns([1, 2], vertical_alignment="center")
    with c1:
        st.markdown(_logo_html(width_px=260, top_pad_px=14), unsafe_allow_html=True)
    with c2:
        st.markdown('<div class="kc-hero"><h1>Payment Summary</h1><p>Secure login required.</p></div>', unsafe_allow_html=True)

    u = _get_secret(["auth", "username"])
    p = _get_secret(["auth", "password"])
    if not u or not p:
        st.error("Missing auth secrets. Add them to Streamlit secrets before using this app.")
        st.stop()

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.form_submit_button("Login"):
            if username == str(u) and password == str(p):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Invalid credentials.")
    st.stop()


def _mixpanel_headers() -> Dict[str, str]:
    auth = _get_secret(["mixpanel", "authorization"])
    if not auth:
        raise RuntimeError("Missing mixpanel.authorization in Streamlit secrets.")
    return {"accept": "text/plain", "authorization": str(auth).strip()}


# =========================
# Core utilities
# =========================
def _excel_safe(v):
    if v is None or v is pd.NA:
        return ""
    if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
        return ""
    if isinstance(v, (np.floating, np.integer)):
        return v.item()
    if isinstance(v, pd.Timestamp):
        if v.tzinfo is not None:
            v = v.tz_convert(None)
        return v.to_pydatetime()
    if isinstance(v, (datetime, date)):
        return v
    if isinstance(v, (list, tuple, set)):
        return ", ".join([str(x) for x in v])
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return v


def _norm_text(val) -> str:
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return str(val).strip().lower()


def _extract_emails(value) -> List[str]:
    if value is None:
        return []
    try:
        if isinstance(value, float) and pd.isna(value):
            return []
    except Exception:
        pass
    found = EMAIL_REGEX.findall(str(value))
    out: List[str] = []
    seen = set()
    for e in found:
        e2 = e.strip().lower()
        if e2 and e2 not in seen:
            seen.add(e2)
            out.append(e2)
    return out


def _pick_first_existing_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    lower_map = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in lower_map:
            return lower_map[c.lower()]
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _parse_time_to_dt(series: pd.Series) -> pd.Series:
    t = pd.to_numeric(series, errors="coerce")
    if t.dropna().empty:
        return pd.to_datetime(series, errors="coerce", utc=True)
    if float(t.median()) > 1e11:  # ms epoch
        t = (t // 1000)
    return pd.to_datetime(t, unit="s", utc=True)


# =========================
# Mixpanel export + DEDUPE
# =========================
def dedupe_mixpanel_export(df: pd.DataFrame) -> pd.DataFrame:
    required = ["event", "distinct_id", "time", "$insert_id"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(f"Missing columns: {missing}. Available: {list(df.columns)}")

    df = df.copy()
    t = pd.to_numeric(df["time"], errors="coerce")
    if t.notna().all():
        if float(t.median()) > 1e11:
            t = (t // 1000)
        df["_time_s"] = t.astype("Int64")
    else:
        dt = pd.to_datetime(df["time"], errors="coerce", utc=True)
        df["_time_s"] = (dt.view("int64") // 10**9).astype("Int64")

    sort_cols = ["_time_s"]
    if "mp_processing_time_ms" in df.columns:
        sort_cols = ["mp_processing_time_ms"] + sort_cols

    df = df.sort_values(sort_cols, kind="mergesort")
    before = len(df)
    df = df.drop_duplicates(subset=["event", "distinct_id", "_time_s", "$insert_id"], keep="last")
    after = len(df)

    df = df.drop(columns=["_time_s"])
    df.attrs["dedupe_removed"] = before - after
    df.attrs["dedupe_before"] = before
    df.attrs["dedupe_after"] = after
    return df


@st.cache_data(show_spinner=False, ttl=600)
def fetch_mixpanel_event_export(project_id: int, base_url: str, from_date: date, to_date: date, event_name: str) -> pd.DataFrame:
    url = f"{base_url.rstrip('/')}/api/2.0/export"
    params = {
        "project_id": int(project_id),
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "event": json.dumps([event_name]),
    }
    resp = requests.get(url, params=params, headers=_mixpanel_headers(), timeout=180)
    if resp.status_code != 200:
        body = (resp.text or "")[:500]
        raise RuntimeError(f"Mixpanel export failed for '{event_name}'. Status {resp.status_code}. Body: {body}")

    objs = []
    for line in resp.text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            objs.append(json.loads(line))
        except json.JSONDecodeError:
            continue

    if not objs:
        return pd.DataFrame()

    raw = pd.DataFrame(objs)
    if "properties" in raw.columns:
        props = pd.json_normalize(raw["properties"])
        raw = pd.concat([raw.drop(columns=["properties"]), props], axis=1)

    if "time" in raw.columns and not raw.empty:
        raw["_dt"] = _parse_time_to_dt(raw["time"])
    return raw


# =========================
# Leads parsing
# =========================
def _split_labels(value) -> List[str]:
    if value is None:
        return []
    try:
        if isinstance(value, float) and pd.isna(value):
            return []
    except Exception:
        pass
    parts = [p.strip() for p in str(value).split(",")]
    return [p for p in parts if p]


def _connected_from_labels(labels: List[str]) -> bool:
    labs = [str(l).strip().lower() for l in (labels or [])]
    if any(l == "not connected" for l in labs):
        return False
    if any("connected" in l for l in labs):
        return True
    return False


def _expand_leads_for_multiple_emails(df: pd.DataFrame, email_cols_priority: List[str]) -> Tuple[pd.DataFrame, List[int]]:
    missing_rows: List[int] = []
    expanded = []

    for i, row in df.iterrows():
        emails: List[str] = []
        for col in email_cols_priority:
            if col in df.columns:
                found = _extract_emails(row[col])
                if found:
                    emails = found
                    break
        if not emails:
            rec = row.to_dict()
            rec["email"] = None
            expanded.append(rec)
            missing_rows.append(i + 2)
            continue
        for e in emails:
            rec = row.to_dict()
            rec["email"] = e
            expanded.append(rec)

    return pd.DataFrame(expanded), missing_rows


# =========================
# Business logic
# =========================
def _filter_credit_excluded(df: pd.DataFrame, text_col: Optional[str]) -> pd.DataFrame:
    if df.empty or not text_col or text_col not in df.columns:
        return df.copy()
    mask = df[text_col].apply(_norm_text).isin(CREDIT_EXCLUDE_DESCS)
    return df[~mask].copy()


def _windowed_email_summary(
    payments_gross: pd.DataFrame,
    refunds_gross: pd.DataFrame,
    payments_ce: pd.DataFrame,
    refunds_ce: pd.DataFrame,
    amount_col: str,
    desc_col: Optional[str],
    refund_amount_col: str,
    days: int = 7,
) -> pd.DataFrame:
    """
    7-day window logic per email.
    """
    d = payments_gross.dropna(subset=["email"]).copy()
    if d.empty:
        cols = [
            "email", "Net_Amount", "Net_Amount_creditExcluded", "Total_Amount",
            "Total_Amount_creditExcluded", "Refund_Amount", "Refund_Amount_creditExcluded",
            "Transactions", "Transactions_creditExcluded", "First_Subscription", "First_Payment_Date"
        ]
        return pd.DataFrame(columns=cols)

    d[amount_col] = pd.to_numeric(d[amount_col], errors="coerce").fillna(0.0)
    d = d.sort_values("_dt", kind="mergesort")

    ce_map = {e: g.sort_values("_dt") for e, g in payments_ce.dropna(subset=["email"]).groupby("email")} if not payments_ce.empty else {}
    ref_map = {e: g.sort_values("_dt") for e, g in refunds_gross.dropna(subset=["email"]).groupby("email")} if not refunds_gross.empty else {}
    ref_ce_map = {e: g.sort_values("_dt") for e, g in refunds_ce.dropna(subset=["email"]).groupby("email")} if not refunds_ce.empty else {}

    out = []
    for email, g in d.groupby("email", sort=False):
        g = g.sort_values("_dt")
        trigger = False
        start = None
        if desc_col and desc_col in g.columns:
            mask = g[desc_col].astype(str).str.contains("Workspace Subscription", case=False, na=False)
            if mask.any():
                trigger = True
                start = g.loc[mask, "_dt"].min()
        if start is None:
            start = g["_dt"].min()

        end = start + timedelta(days=days)

        gross_total = float(g[(g["_dt"] >= start) & (g["_dt"] <= end)][amount_col].sum())
        gross_txn = int(g[(g["_dt"] >= start) & (g["_dt"] <= end)].shape[0])

        g_ce = ce_map.get(email)
        ce_total = float(g_ce[(g_ce["_dt"] >= start) & (g_ce["_dt"] <= end)][amount_col].sum()) if g_ce is not None else 0.0
        ce_txn = int(g_ce[(g_ce["_dt"] >= start) & (g_ce["_dt"] <= end)].shape[0]) if g_ce is not None else 0

        g_ref = ref_map.get(email)
        ref_total = float(g_ref[(g_ref["_dt"] >= start) & (g_ref["_dt"] <= end)][refund_amount_col].sum()) if g_ref is not None else 0.0

        g_ref_ce = ref_ce_map.get(email)
        ref_ce_total = float(g_ref_ce[(g_ref_ce["_dt"] >= start) & (g_ref_ce["_dt"] <= end)][refund_amount_col].sum()) if g_ref_ce is not None else 0.0

        out.append({
            "email": email,
            "Net_Amount": gross_total - ref_total,
            "Net_Amount_creditExcluded": ce_total - ref_ce_total,
            "Total_Amount": gross_total,
            "Total_Amount_creditExcluded": ce_total,
            "Refund_Amount": ref_total,
            "Refund_Amount_creditExcluded": ref_ce_total,
            "Transactions": gross_txn,
            "Transactions_creditExcluded": ce_txn,
            "First_Subscription": "TRUE" if trigger else "FALSE",
            "First_Payment_Date": pd.to_datetime(start, utc=True).tz_convert(None) if pd.notna(start) else None,
        })
    return pd.DataFrame(out)


def _strict_range_email_summary(
    payments_gross: pd.DataFrame,
    refunds_gross: pd.DataFrame,
    payments_ce: pd.DataFrame,
    refunds_ce: pd.DataFrame,
    amount_col: str,
    refund_amount_col: str,
) -> pd.DataFrame:
    """
    Strict selected range summation (Period Revenue) per email.
    """
    def get_sums(df, group_col, sum_col, out_col):
        if df.empty:
            return pd.Series(dtype=float)
        return df.groupby(group_col)[sum_col].sum().rename(out_col)

    p_gross = get_sums(payments_gross, "email", amount_col, "Period_Total_Amount")
    p_ce = get_sums(payments_ce, "email", amount_col, "Period_Total_Amount_creditExcluded")
    r_gross = get_sums(refunds_gross, "email", refund_amount_col, "Period_Refund_Amount")
    r_ce = get_sums(refunds_ce, "email", refund_amount_col, "Period_Refund_Amount_creditExcluded")

    df = pd.concat([p_gross, p_ce, r_gross, r_ce], axis=1).fillna(0.0)
    df.index.name = "email"
    df = df.reset_index()

    df["Period_Net_Amount"] = df["Period_Total_Amount"] - df["Period_Refund_Amount"]
    df["Period_Net_Amount_creditExcluded"] = df["Period_Total_Amount_creditExcluded"] - df["Period_Refund_Amount_creditExcluded"]

    return df


def _add_totals_row(df: pd.DataFrame, label_col: Optional[str] = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    num_cols = out.select_dtypes(include=[np.number]).columns.tolist()
    totals = {c: float(pd.to_numeric(out[c], errors="coerce").fillna(0).sum()) for c in num_cols}
    row = {c: "" for c in out.columns}
    if label_col and label_col in out.columns:
        row[label_col] = "TOTAL"
    else:
        row[out.columns[0]] = "TOTAL"
    row.update(totals)
    return pd.concat([out, pd.DataFrame([row])], ignore_index=True)


def _style_totals_row(df: pd.DataFrame):
    if df is None or df.empty:
        return df

    def _row_style(row):
        is_total = str(row.iloc[0]).strip().upper() == "TOTAL"
        return ["font-weight: bold" if is_total else "" for _ in row]

    return df.style.apply(_row_style, axis=1)


def _style_sheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        first = row[0].value
        if first and str(first).strip().upper() == "TOTAL":
            for cell in row:
                cell.font = Font(bold=True)
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col[:250]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 50)


def _build_excel(
    leads_with_payments: pd.DataFrame,
    leads_nonzero: pd.DataFrame,
    email_summary: pd.DataFrame,
    owner_summary: pd.DataFrame,
    owner_x_connected: pd.DataFrame,
    connected_summary: pd.DataFrame,
    label_summary: pd.DataFrame,
    time_summary: pd.DataFrame,
    hour_summary: pd.DataFrame,
    duplicate_leads: pd.DataFrame,
    self_converted_emails: pd.DataFrame,
    overall_metrics_df: pd.DataFrame,
    logs_df: pd.DataFrame,
    fig_owner,
    fig_owner_conn,
    fig_time,
    from_date: date,
    to_date: date,
) -> Tuple[str, bytes]:
    wb = Workbook()

    def add_sheet(title: str, df: pd.DataFrame, label_col: Optional[str] = None, add_totals: bool = True):
        ws = wb.create_sheet(title)
        df2 = _add_totals_row(df, label_col=label_col) if add_totals and df is not None and not df.empty else df
        if df2 is None or df2.empty:
            ws.append(["No data"])
            _style_sheet(ws)
            return ws
        for r in dataframe_to_rows(df2, index=False, header=True):
            ws.append([_excel_safe(x) for x in r])
        _style_sheet(ws)
        return ws

    ws0 = wb.active
    ws0.title = "Overall Metrics"
    if not overall_metrics_df.empty:
        for r in dataframe_to_rows(overall_metrics_df, index=False, header=True):
            ws0.append([_excel_safe(x) for x in r])
    _style_sheet(ws0)

    add_sheet("Leads_with_Payments", leads_with_payments, label_col=None)
    add_sheet("Leads_Payments_NonZero", leads_nonzero, label_col=None)
    add_sheet("Email_Summary", email_summary, label_col="email")
    add_sheet("Owner_Summary", owner_summary, label_col=owner_summary.columns[0] if not owner_summary.empty else None, add_totals=False)
    add_sheet("Owner_x_Connected", owner_x_connected, label_col=owner_x_connected.columns[0] if not owner_x_connected.empty else None)
    add_sheet("Connected_Summary", connected_summary, label_col="Connected" if "Connected" in connected_summary.columns else None)
    add_sheet("Label_Summary", label_summary, label_col="Label" if "Label" in label_summary.columns else None)
    add_sheet("Time_Summary", time_summary, label_col=time_summary.columns[0] if not time_summary.empty else None)
    add_sheet("Hour_Summary", hour_summary, label_col="Lead_Created_Hour" if "Lead_Created_Hour" in hour_summary.columns else None)
    add_sheet("Duplicate_Leads_By_Email", duplicate_leads, label_col="email" if "email" in duplicate_leads.columns else None)
    add_sheet("SelfConverted_Emails", self_converted_emails, label_col="email" if "email" in self_converted_emails.columns else None)
    add_sheet("Logs", logs_df, label_col=None)

    ws_chart = wb.create_sheet("Charts")

    def add_fig(ws, fig, anchor, table_df: pd.DataFrame, table_anchor_col: int, table_anchor_row: int):
        img_bytes = BytesIO()
        fig.savefig(img_bytes, format="png", dpi=150, bbox_inches="tight")
        img_bytes.seek(0)
        img = XLImage(img_bytes)
        img.anchor = anchor
        ws.add_image(img)
        if table_df is not None and not table_df.empty:
            rounded = table_df.copy()
            for c in rounded.select_dtypes(include=[np.number]).columns:
                rounded[c] = pd.to_numeric(rounded[c], errors="coerce").fillna(0).round(0).astype(int)
            for r_idx, row in enumerate(dataframe_to_rows(rounded, index=False, header=True), table_anchor_row):
                for c_idx, value in enumerate(row, table_anchor_col):
                    ws.cell(row=r_idx, column=c_idx, value=_excel_safe(value))

    add_fig(ws_chart, fig_owner, "A1", owner_summary, 8, 1)
    add_fig(ws_chart, fig_owner_conn, "A26", owner_x_connected, 8, 26)
    add_fig(ws_chart, fig_time, "A51", time_summary, 8, 51)
    _style_sheet(ws_chart)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    out = BytesIO()
    wb.save(out)
    fname = f"payment_summary_{from_date.strftime('%b%d').lower()}_{to_date.strftime('%b%d').lower()}.xlsx"
    return fname, out.getvalue()


# =========================
# Main app
# =========================
def main():
    st.set_page_config(page_title="KrispCall Payment Summary", page_icon="📈", layout="wide")
    require_login()
    _inject_brand_css()

    st.markdown('<div style="height:10px;"></div>', unsafe_allow_html=True)
    l, r = st.columns([1, 3], vertical_alignment="center")
    with l:
        st.markdown(_logo_html(width_px=240, top_pad_px=14), unsafe_allow_html=True)
    with r:
        st.markdown(
            '<div class="kc-hero"><h1>KrispCall Payment Summary</h1><p>Leads reconciliation with Mixpanel transactions</p></div>',
            unsafe_allow_html=True,
        )

    with st.sidebar:
        st.markdown("### Date Selection")
        today = date.today()
        first_of_month = today.replace(day=1)
        default_start = first_of_month - timedelta(days=4)
        default_end = today - timedelta(days=1)
        from_date = st.date_input("Date from", value=default_start)
        to_date = st.date_input("Date to", value=default_end)
        st.markdown("---")
        st.markdown("### Rules")

    st.markdown('<div class="kc-card">', unsafe_allow_html=True)
    leads_file = st.file_uploader("Upload Leads CSV", type=["csv"])
    run = st.button("Run Analysis", type="primary", disabled=(leads_file is None))
    st.markdown("</div>", unsafe_allow_html=True)

    if not run:
        st.stop()

    logs: List[str] = []

    with st.spinner("Running analysis..."):
        leads_raw = pd.read_csv(leads_file)

        owner_col = _pick_first_existing_column(leads_raw, ["Lead - Owner", "Deal - Owner", "Owner", "owner"]) or "Owner"
        created_col = _pick_first_existing_column(leads_raw, ["Lead - Lead created on", "Lead created on", "Created on"])
        label_col = _pick_first_existing_column(leads_raw, ["Lead - Label", "Label", "Labels", "Lead - Labels"])

        email_cols_priority: List[str] = []
        for cand in ["Person - Email", "Lead - User Email"]:
            col = _pick_first_existing_column(leads_raw, [cand])
            if col and col not in email_cols_priority:
                email_cols_priority.append(col)
        if not email_cols_priority:
            fallback = [c for c in leads_raw.columns if "email" in c.lower()]
            email_cols_priority = fallback[:5]

        expanded_leads, missing_rows = _expand_leads_for_multiple_emails(leads_raw, email_cols_priority)
        if missing_rows:
            logs.append(f"Missing email for {len(missing_rows)} lead row(s).")

        if label_col and label_col in expanded_leads.columns:
            expanded_leads["labels_list"] = expanded_leads[label_col].apply(_split_labels)
        else:
            expanded_leads["labels_list"] = [[] for _ in range(len(expanded_leads))]

        expanded_leads["Connected"] = expanded_leads["labels_list"].apply(_connected_from_labels)

        if created_col and created_col in expanded_leads.columns:
            expanded_leads["_lead_created_dt"] = pd.to_datetime(expanded_leads[created_col], errors="coerce")
        else:
            expanded_leads["_lead_created_dt"] = pd.NaT

        expanded_leads["email"] = expanded_leads["email"].apply(
            lambda x: str(x).strip().lower() if x is not None and str(x).strip() != "" else None
        )

        # Lead Count metrics base (before dedup). Exclude India and Junk Lead for Leads Generated.
        country_col = _pick_first_existing_column(leads_raw, ["Person - Country", "Country", "Person Country"])
        labels_raw_col = label_col

        lead_counts_base = expanded_leads.copy()

        mask_india = pd.Series(False, index=lead_counts_base.index)
        if country_col and country_col in lead_counts_base.columns:
            mask_india = lead_counts_base[country_col].apply(_norm_text).eq("india")

        mask_junk = pd.Series(False, index=lead_counts_base.index)
        if labels_raw_col and labels_raw_col in lead_counts_base.columns:
            mask_junk = lead_counts_base[labels_raw_col].astype(str).str.contains("junk lead", case=False, na=False)

        leads_generated_df = lead_counts_base[~mask_india & ~mask_junk].copy()
        leads_attempted_df = leads_generated_df[~leads_generated_df[owner_col].apply(_norm_text).eq(EXCLUDED_OWNER_CANON)].copy()
        leads_connected_df = leads_attempted_df[leads_attempted_df["Connected"] == True].copy()

        metric_leads_generated = int(len(leads_generated_df))
        metric_leads_attempted = int(len(leads_attempted_df))
        metric_leads_connected = int(len(leads_connected_df))

        lead_count_relation_ok = (metric_leads_connected <= metric_leads_attempted <= metric_leads_generated)
        if not lead_count_relation_ok:
            logs.append(
                f"AUDIT FAIL: Lead counts relationship unexpected. Generated={metric_leads_generated}, "
                f"Attempted={metric_leads_attempted}, Connected={metric_leads_connected}"
            )

        # Mixpanel
        pid = int(_get_secret(["mixpanel", "project_id"]))
        base = _get_secret(["mixpanel", "base_url"], "https://data-eu.mixpanel.com")

        payments_raw = fetch_mixpanel_event_export(pid, base, from_date, to_date, "New Payment Made")
        refunds_raw = fetch_mixpanel_event_export(pid, base, from_date, to_date, "Refund Granted")

        if not payments_raw.empty:
            payments = dedupe_mixpanel_export(payments_raw)
            logs.append(
                f"Payments raw rows: {payments.attrs.get('dedupe_before', len(payments_raw))}. "
                f"Dedupe removed: {payments.attrs.get('dedupe_removed', 0)}."
            )
        else:
            payments = payments_raw.copy()
            logs.append("Payments raw rows: 0.")

        if not refunds_raw.empty:
            refunds = dedupe_mixpanel_export(refunds_raw)
            logs.append(
                f"Refunds raw rows: {refunds.attrs.get('dedupe_before', len(refunds_raw))}. "
                f"Dedupe removed: {refunds.attrs.get('dedupe_removed', 0)}."
            )
        else:
            refunds = refunds_raw.copy()
            logs.append("Refunds raw rows: 0.")

        pay_email_col = _pick_first_existing_column(payments, ["$email", "email", "Email", "EMAIL", "User Email", "user.email"])
        ref_email_col = _pick_first_existing_column(refunds, ["User Email", "user.email", "$email", "email", "Email", "EMAIL"])

        payments["email"] = (
            payments[pay_email_col].apply(lambda v: (_extract_emails(v)[0] if _extract_emails(v) else None))
            if pay_email_col and pay_email_col in payments.columns
            else None
        )
        refunds["email"] = (
            refunds[ref_email_col].apply(lambda v: (_extract_emails(v)[0] if _extract_emails(v) else None))
            if ref_email_col and ref_email_col in refunds.columns
            else None
        )

        if "email" in payments.columns:
            payments["email"] = payments["email"].apply(lambda x: str(x).strip().lower() if x is not None and str(x).strip() != "" else None)
        if "email" in refunds.columns:
            refunds["email"] = refunds["email"].apply(lambda x: str(x).strip().lower() if x is not None and str(x).strip() != "" else None)

        amount_col = _pick_first_existing_column(payments, ["Amount", "amount", "Amount Paid"])
        desc_col = _pick_first_existing_column(payments, ["Amount Description", "description", "Plan"])
        refund_amount_col = _pick_first_existing_column(refunds, ["Refund Amount", "refund_amount", "Amount", "amount"])
        refund_desc_col = _pick_first_existing_column(refunds, ["Refunded Transaction description", "Refunded Transaction Description", "Refunded Transaction"])

        if not amount_col:
            raise RuntimeError("Could not find payment amount column.")
        if not refund_amount_col:
            refunds["Refund Amount"] = 0.0
            refund_amount_col = "Refund Amount"

        payments[amount_col] = pd.to_numeric(payments[amount_col], errors="coerce").fillna(0.0)
        refunds[refund_amount_col] = pd.to_numeric(refunds[refund_amount_col], errors="coerce").fillna(0.0)

        payments_all_ce = _filter_credit_excluded(payments, desc_col)
        refunds_all_ce = _filter_credit_excluded(refunds, refund_desc_col)

        # Email Summaries
        lead_emails = set(expanded_leads["email"].dropna().unique())

        payments_leads = payments[payments["email"].isin(lead_emails)].copy()
        refunds_leads = refunds[refunds["email"].isin(lead_emails)].copy()
        payments_leads_ce = payments_all_ce[payments_all_ce["email"].isin(lead_emails)].copy()
        refunds_leads_ce = refunds_all_ce[refunds_all_ce["email"].isin(lead_emails)].copy()

        email_summary_7d = _windowed_email_summary(
            payments_gross=payments_leads,
            refunds_gross=refunds_leads,
            payments_ce=payments_leads_ce,
            refunds_ce=refunds_leads_ce,
            amount_col=amount_col,
            desc_col=desc_col,
            refund_amount_col=refund_amount_col,
            days=7,
        )

        email_summary_period = _strict_range_email_summary(
            payments_gross=payments,
            refunds_gross=refunds,
            payments_ce=payments_all_ce,
            refunds_ce=refunds_all_ce,
            amount_col=amount_col,
            refund_amount_col=refund_amount_col,
        )

        joined = expanded_leads.merge(email_summary_7d, on="email", how="left")
        joined = joined.merge(email_summary_period, on="email", how="left")

        numeric_cols_7d = [
            "Net_Amount",
            "Net_Amount_creditExcluded",
            "Total_Amount",
            "Total_Amount_creditExcluded",
            "Refund_Amount",
            "Refund_Amount_creditExcluded",
        ]
        numeric_cols_period = [
            "Period_Net_Amount",
            "Period_Net_Amount_creditExcluded",
            "Period_Total_Amount",
            "Period_Total_Amount_creditExcluded",
            "Period_Refund_Amount",
            "Period_Refund_Amount_creditExcluded",
        ]
        for c in numeric_cols_7d + numeric_cols_period:
            joined[c] = pd.to_numeric(joined[c], errors="coerce").fillna(0.0)

        # Bases for Summaries (Exclude Pipedrive KrispCall)
        summ_base = joined[~joined[owner_col].apply(_norm_text).eq(EXCLUDED_OWNER_CANON)].copy()
        summ_base["_lead_created_dt"] = pd.to_datetime(summ_base["_lead_created_dt"], errors="coerce")
        summ_dedup = (
            summ_base.sort_values(["email", "_lead_created_dt"], kind="mergesort")
            .drop_duplicates(subset=["email"], keep="first")
            .copy()
        )

        # Overall Revenue (Period)
        overall_rev_sum = float(payments[amount_col].sum()) if not payments.empty else 0.0
        overall_ref_sum = float(refunds[refund_amount_col].sum()) if not refunds.empty else 0.0
        metric_overall_revenue = float(overall_rev_sum - overall_ref_sum)

        # Overall Conversion set (Workspace Subscription payments only)
        sub_mask = pd.Series(False, index=payments.index)
        if desc_col and desc_col in payments.columns:
            sub_mask = payments[desc_col].astype(str).str.contains("workspace subscription", case=False, na=False)

        overall_conversion_emails = set(
            payments.loc[sub_mask, "email"].dropna().astype(str).str.strip().str.lower().unique()
        )
        metric_overall_conversions = int(len(overall_conversion_emails))

        # Lead dedup mapping (email -> first owner across ALL leads, including Pipedrive)
        lead_dedup_all = expanded_leads.dropna(subset=["email"]).copy()
        lead_dedup_all["email"] = lead_dedup_all["email"].astype(str).str.strip().str.lower()
        lead_dedup_all["_lead_sort_dt"] = pd.to_datetime(lead_dedup_all["_lead_created_dt"], errors="coerce").fillna(
            pd.Timestamp("1970-01-01")
        )
        lead_dedup_all = (
            lead_dedup_all.sort_values(["email", "_lead_sort_dt"], kind="mergesort")
            .drop_duplicates(subset=["email"], keep="first")
        )
        lead_dedup_all["_owner_norm"] = lead_dedup_all[owner_col].apply(_norm_text)
        lead_first_owner = lead_dedup_all.set_index("email")["_owner_norm"].to_dict()

        # Self Converted definition
        self_converted_emails_set = {
            e for e in overall_conversion_emails
            if (e not in lead_first_owner) or (lead_first_owner.get(e) == EXCLUDED_OWNER_CANON)
        }
        sales_conversion_emails_set = overall_conversion_emails - self_converted_emails_set

        self_converted_emails = sorted(list(self_converted_emails_set))
        sales_conversion_emails = sorted(list(sales_conversion_emails_set))

        self_converted_count = int(len(self_converted_emails_set))
        sales_conversions_count = int(len(sales_conversion_emails_set))

        conversion_partition_ok = (metric_overall_conversions == (self_converted_count + sales_conversions_count))
        if not conversion_partition_ok:
            logs.append(
                f"AUDIT FAIL: Overall Conversion {metric_overall_conversions} != "
                f"Sales Conversions {sales_conversions_count} + Self Converted {self_converted_count}"
            )

        # Sales Effort Conversion emails (Sales conversions that also appear in Connected leads excluding Pipedrive)
        connected_emails_set = set(
            summ_base.loc[summ_base["Connected"] == True, "email"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.lower()
            .unique()
        )
        sales_effort_conversion_emails_set = sales_conversion_emails_set.intersection(connected_emails_set)
        sales_effort_conversions_count = int(len(sales_effort_conversion_emails_set))

        effort_subset_ok = (sales_effort_conversions_count <= sales_conversions_count)
        if not effort_subset_ok:
            logs.append(
                f"AUDIT FAIL: Sales Effort Conversions {sales_effort_conversions_count} > Sales Conversions {sales_conversions_count}"
            )

        # Segment revenue helpers
        def calc_segment_metrics(emails_list: List[str]):
            if not emails_list:
                return 0.0, 0.0

            p_pay = payments[payments["email"].isin(emails_list)]
            p_ref = refunds[refunds["email"].isin(emails_list)]
            period_net = float(p_pay[amount_col].sum() - p_ref[refund_amount_col].sum())

            p_pay_ce = payments_all_ce[payments_all_ce["email"].isin(emails_list)]
            p_ref_ce = refunds_all_ce[refunds_all_ce["email"].isin(emails_list)]

            summ = _windowed_email_summary(
                payments_gross=p_pay,
                refunds_gross=p_ref,
                payments_ce=p_pay_ce,
                refunds_ce=p_ref_ce,
                amount_col=amount_col,
                desc_col=desc_col,
                refund_amount_col=refund_amount_col,
                days=7,
            )
            net_7d = float(summ["Net_Amount"].sum()) if not summ.empty else 0.0
            return period_net, net_7d

        sc_period_net, sc_7d_net = calc_segment_metrics(list(self_converted_emails_set))
        sales_conv_period_net, sales_conv_7d_net = calc_segment_metrics(list(sales_conversion_emails_set))
        sales_effort_period_net, sales_effort_7d_net = calc_segment_metrics(list(sales_effort_conversion_emails_set))

        # Overall Metrics table
        overall_metrics_data = [
            {"Group": "Overall", "Metric": "Overall Revenue (Period)", "Value": metric_overall_revenue, "Description": "Payments minus refunds in selected range"},
            {"Group": "Overall", "Metric": "Overall Conversion", "Value": metric_overall_conversions, "Description": "Unique emails with Workspace Subscription payment"},

            {"Group": "Self Converted", "Metric": "Self-Converted Count (Conversion Count)", "Value": self_converted_count, "Description": "Workspace Subscription emails with no lead, or deduped first lead owner is Pipedrive KrispCall"},
            {"Group": "Self Converted", "Metric": "Self-Converted Net Revenue (Whole Period)", "Value": sc_period_net, "Description": "Net revenue in range for self converted conversion emails"},
            {"Group": "Self Converted", "Metric": "Self-Converted Net Revenue (7 day)", "Value": sc_7d_net, "Description": "7-day window net revenue for self converted conversion emails"},

            {"Group": "Sales Conversions", "Metric": "Sales Conversions Count", "Value": sales_conversions_count, "Description": "Overall Conversion minus Self Converted, audited"},
            {"Group": "Sales Conversions", "Metric": "Sales Effort Conversions Count", "Value": sales_effort_conversions_count, "Description": "Sales Conversion emails that appear in Connected leads (excluding Pipedrive)"},
            {"Group": "Sales Conversions", "Metric": "Sales Conversion Revenue (Whole Period)", "Value": sales_conv_period_net, "Description": "Net revenue in range for sales conversion emails"},
            {"Group": "Sales Conversions", "Metric": "Sales Effort Revenue (Whole Period)", "Value": sales_effort_period_net, "Description": "Net revenue in range for sales effort conversion emails"},
            {"Group": "Sales Conversions", "Metric": "Sales Conversion Revenue (7 day)", "Value": sales_conv_7d_net, "Description": "7-day window net for sales conversion emails"},
            {"Group": "Sales Conversions", "Metric": "Sales Effort Revenue (7 day)", "Value": sales_effort_7d_net, "Description": "7-day window net for sales effort conversion emails"},

            {"Group": "Lead Count", "Metric": "Leads Generated", "Value": metric_leads_generated, "Description": 'Lead rows excluding Person - Country "India" and Lead - Label containing "Junk Lead". Counted before dedup'},
            {"Group": "Lead Count", "Metric": "Leads Attempted", "Value": metric_leads_attempted, "Description": "Leads Generated excluding owner Pipedrive KrispCall. Counted before dedup"},
            {"Group": "Lead Count", "Metric": "Lead Connected", "Value": metric_leads_connected, "Description": "Connected leads excluding owner Pipedrive KrispCall. Counted before dedup"},
        ]
        overall_metrics_df = pd.DataFrame(overall_metrics_data)

        # Summaries Construction
        summ_cols = numeric_cols_7d + numeric_cols_period

        def build_summary(group_cols: List[str], df_base: pd.DataFrame, df_dedup: pd.DataFrame):
            rev = df_dedup.groupby(group_cols, as_index=False)[summ_cols].sum()
            counts = df_base.groupby(group_cols, as_index=False).size().rename(columns={"size": "Lead_Count"})

            pay_mask = (df_base["Period_Total_Amount"] > 0) | (df_base["Total_Amount"] > 0)
            payers = (
                df_base[pay_mask]
                .groupby(group_cols, as_index=False)["email"]
                .nunique()
                .rename(columns={"email": "Paying_Users"})
            )

            final = rev.merge(counts, on=group_cols, how="outer")
            final = final.merge(payers, on=group_cols, how="outer")
            final = final.fillna(0)

            if "Period_Net_Amount" in final.columns:
                final = final.sort_values("Period_Net_Amount", ascending=False)
            return final

        owner_summary = build_summary([owner_col], summ_base, summ_dedup)
        connected_summary = build_summary(["Connected"], summ_base, summ_dedup)
        owner_x_connected = build_summary([owner_col, "Connected"], summ_base, summ_dedup)

        labels_base = summ_base.explode("labels_list").rename(columns={"labels_list": "Label"})
        labels_base["Label"] = labels_base["Label"].fillna("").astype(str).str.strip()
        labels_base = labels_base[labels_base["Label"] != ""].copy()
        labels_base["Connected_Label"] = labels_base["Label"].str.lower().apply(
            lambda s: False if s.strip() == "not connected" else ("connected" in s)
        )

        labels_dedup = summ_dedup.explode("labels_list").rename(columns={"labels_list": "Label"})
        labels_dedup["Label"] = labels_dedup["Label"].fillna("").astype(str).str.strip()
        labels_dedup = labels_dedup[labels_dedup["Label"] != ""].copy()
        labels_dedup["Connected_Label"] = labels_dedup["Label"].str.lower().apply(
            lambda s: False if s.strip() == "not connected" else ("connected" in s)
        )

        label_summary = build_summary(["Label", "Connected_Label"], labels_base, labels_dedup)

        summ_base["Lead_Created_Date"] = summ_base["_lead_created_dt"].dt.date
        summ_dedup["Lead_Created_Date"] = summ_dedup["_lead_created_dt"].dt.date
        time_summary = build_summary(["Lead_Created_Date"], summ_base, summ_dedup).sort_values("Lead_Created_Date")

        # Hour of day summary
        summ_base["Lead_Created_Hour"] = pd.to_datetime(summ_base["_lead_created_dt"], errors="coerce").dt.hour
        summ_dedup["Lead_Created_Hour"] = pd.to_datetime(summ_dedup["_lead_created_dt"], errors="coerce").dt.hour
        hour_summary = build_summary(["Lead_Created_Hour"], summ_base, summ_dedup).copy()

        def _hour_label(v):
            try:
                if pd.isna(v):
                    return "Unknown"
                h = int(v)
                if h < 0 or h > 23:
                    return "Unknown"
                return f"{h:02d}:00"
            except Exception:
                return "Unknown"

        hour_summary["Lead_Created_Hour"] = hour_summary["Lead_Created_Hour"].apply(_hour_label)
        hour_summary["_sort"] = hour_summary["Lead_Created_Hour"].apply(
            lambda s: 99 if s == "Unknown" else int(str(s).split(":")[0])
        )
        hour_summary = hour_summary.sort_values("_sort").drop(columns=["_sort"])

        dup_mask = summ_base["email"].notna() & summ_base["email"].duplicated(keep=False)
        duplicate_leads = summ_base.loc[dup_mask].sort_values(["email", "_lead_created_dt"], kind="mergesort").copy()

        # Self Converted detail table (new definition)
        pay_sc = payments[payments["email"].isin(self_converted_emails_set)]
        ref_sc = refunds[refunds["email"].isin(self_converted_emails_set)]
        pay_sc_ce = payments_all_ce[payments_all_ce["email"].isin(self_converted_emails_set)]
        ref_sc_ce = refunds_all_ce[refunds_all_ce["email"].isin(self_converted_emails_set)]

        sc_summ_7d = _windowed_email_summary(pay_sc, ref_sc, pay_sc_ce, ref_sc_ce, amount_col, desc_col, refund_amount_col, 7)
        sc_summ_per = _strict_range_email_summary(pay_sc, ref_sc, pay_sc_ce, ref_sc_ce, amount_col, refund_amount_col)
        self_converted_fact = (
            sc_summ_7d.merge(sc_summ_per, on="email", how="left")
            .sort_values("Period_Net_Amount", ascending=False)
        )

        # Charts
        fig_owner, ax_owner = plt.subplots(figsize=(10, 6))
        if not owner_summary.empty:
            chart_df = owner_summary.set_index(owner_col)[["Period_Net_Amount", "Net_Amount"]]
            chart_df.columns = ["Strict Range Net", "7-Day Window Net"]
            chart_df.plot(kind="bar", ax=ax_owner)
        ax_owner.set_title("Net Revenue by Owner")
        plt.tight_layout()

        fig_owner_conn, ax_owner_conn = plt.subplots(figsize=(10, 6))
        if not owner_x_connected.empty:
            pivot = (
                owner_x_connected.pivot_table(index=owner_col, columns="Connected", values="Period_Net_Amount", aggfunc="sum")
                .fillna(0.0)
            )
            pivot.plot(kind="bar", stacked=True, ax=ax_owner_conn)
        ax_owner_conn.set_title("Net Revenue (Strict Range) by Owner and Connected")
        plt.tight_layout()

        fig_time, ax_time = plt.subplots(figsize=(10, 5))
        if not time_summary.empty:
            ax_time.plot(pd.to_datetime(time_summary["Lead_Created_Date"]), time_summary["Period_Net_Amount"], label="Strict Range")
            ax_time.plot(pd.to_datetime(time_summary["Lead_Created_Date"]), time_summary["Net_Amount"], label="7-Day Window", linestyle="--")
            ax_time.legend()
        ax_time.set_title("Net Amount by Lead Created Date")
        plt.tight_layout()

        # Excel export tables
        joined_export = joined.copy()
        joined_export["labels_list"] = joined_export["labels_list"].apply(lambda x: ", ".join(x) if isinstance(x, list) else "")
        joined_nonzero = joined[(joined["Total_Amount"].fillna(0).ne(0)) | (joined["Period_Total_Amount"].fillna(0).ne(0))].copy()
        joined_nonzero_export = joined_nonzero.copy()
        joined_nonzero_export["labels_list"] = joined_nonzero_export["labels_list"].apply(lambda x: ", ".join(x) if isinstance(x, list) else "")

        logs_df = pd.DataFrame({"log": logs})

        excel_name, excel_bytes = _build_excel(
            leads_with_payments=joined_export,
            leads_nonzero=joined_nonzero_export,
            email_summary=joined,
            owner_summary=owner_summary,
            owner_x_connected=owner_x_connected,
            connected_summary=connected_summary,
            label_summary=label_summary,
            time_summary=time_summary,
            hour_summary=hour_summary,
            duplicate_leads=duplicate_leads,
            self_converted_emails=self_converted_fact,
            overall_metrics_df=overall_metrics_df,
            logs_df=logs_df,
            fig_owner=fig_owner,
            fig_owner_conn=fig_owner_conn,
            fig_time=fig_time,
            from_date=from_date,
            to_date=to_date,
        )

        # =========================
        # AUDIT DATAFRAMES (for Audit tab)
        # =========================
        audit_summary_df = pd.DataFrame(
            [
                {"Check": "Conversion Partition", "Status": "PASS" if conversion_partition_ok else "FAIL", "Notes": "Overall Conversion equals Sales Conversions plus Self Converted"},
                {"Check": "Effort Subset", "Status": "PASS" if effort_subset_ok else "FAIL", "Notes": "Sales Effort Conversions is subset of Sales Conversions"},
                {"Check": "Lead Count Ordering", "Status": "PASS" if lead_count_relation_ok else "FAIL", "Notes": "Lead Connected <= Leads Attempted <= Leads Generated"},
            ]
        )

        audit_counts_df = pd.DataFrame(
            [
                {"Metric": "Overall Conversion emails", "Value": metric_overall_conversions},
                {"Metric": "Self Converted emails", "Value": self_converted_count},
                {"Metric": "Sales Conversion emails", "Value": sales_conversions_count},
                {"Metric": "Sales Effort Conversion emails", "Value": sales_effort_conversions_count},
                {"Metric": "Leads Generated (rows)", "Value": metric_leads_generated},
                {"Metric": "Leads Attempted (rows)", "Value": metric_leads_attempted},
                {"Metric": "Lead Connected (rows)", "Value": metric_leads_connected},
            ]
        )

        def _segment_for_email(e: str) -> str:
            if e in self_converted_emails_set:
                return "Self Converted"
            if e in sales_conversion_emails_set:
                return "Sales Conversion"
            return "Unclassified"

        # Classification table for conversion emails
        conv_list = sorted(list(overall_conversion_emails))
        conv_owner = [lead_first_owner.get(e, "(no lead)") for e in conv_list]
        conv_has_any_lead = [e in lead_first_owner for e in conv_list]
        conv_in_non_pipedrive_leads = [e in set(summ_base["email"].dropna().unique()) for e in conv_list]
        conv_connected_in_non_pipedrive = [e in connected_emails_set for e in conv_list]
        conv_segment = [_segment_for_email(e) for e in conv_list]

        conversion_classification_df = pd.DataFrame(
            {
                "email": conv_list,
                "segment": conv_segment,
                "first_lead_owner_dedup": conv_owner,
                "has_any_lead": conv_has_any_lead,
                "has_non_pipedrive_lead": conv_in_non_pipedrive_leads,
                "connected_in_non_pipedrive_leads": conv_connected_in_non_pipedrive,
            }
        )

        # Owner distribution for conversion emails
        owner_dist_df = (
            conversion_classification_df["first_lead_owner_dedup"]
            .value_counts()
            .rename_axis("first_lead_owner_dedup")
            .reset_index(name="conversion_emails")
        )

        # Small lists for quick view
        def _head_list(s: List[str], n: int = 50) -> pd.DataFrame:
            return pd.DataFrame({"email": s[:n]})

        self_conv_preview_df = _head_list(self_converted_emails, 50)
        sales_conv_preview_df = _head_list(sales_conversion_emails, 50)
        effort_conv_preview_df = _head_list(sorted(list(sales_effort_conversion_emails_set)), 50)

        audit_lines = [x for x in logs if str(x).startswith("AUDIT")]

        # Download for audit classification
        audit_csv_bytes = conversion_classification_df.to_csv(index=False).encode("utf-8")

    # =========================
    # UI TABS
    # =========================
    (
        tab_overall,
        tab_overview,
        tab_tables,
        tab_summaries,
        tab_time,
        tab_hour,
        tab_audit,
        tab_export,
        tab_logs,
    ) = st.tabs(
        ["Overall Metrics", "Overview", "Main Tables", "Summaries", "Time", "Hour of Day", "Audit", "Export", "Logs"]
    )

    with tab_overall:
        st.markdown("### Overall Business Metrics (Selected Date Range)")
        st.dataframe(overall_metrics_df, use_container_width=True)

        st.markdown("---")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Overall Revenue (Period)", f"{metric_overall_revenue:,.2f}")
        c2.metric("Overall Conversion", f"{metric_overall_conversions:,}")
        c3.metric("Sales Conversions Count", f"{sales_conversions_count:,}")
        c4.metric("Self-Converted Count", f"{self_converted_count:,}")

    with tab_overview:
        total_net_period = float(pd.to_numeric(owner_summary["Period_Net_Amount"], errors="coerce").fillna(0).sum()) if not owner_summary.empty else 0.0
        total_net_7d = float(pd.to_numeric(owner_summary["Net_Amount"], errors="coerce").fillna(0).sum()) if not owner_summary.empty else 0.0

        c1, c2, c3 = st.columns(3)
        c1.metric("Sum of Owners (Period Net)", f"{total_net_period:,.2f}", help="Sum of Period Net for all owners (excluding Pipedrive KrispCall)")
        c2.metric("Sum of Owners (7-Day Net)", f"{total_net_7d:,.2f}", help="Sum of 7-day window Net for all owners (excluding Pipedrive KrispCall)")
        c3.metric("Self-Converted Count", f"{self_converted_count:,}")

        st.pyplot(fig_owner, use_container_width=True)

    with tab_tables:
        st.markdown("#### Leads with payments")
        st.dataframe(joined_export, use_container_width=True)
        st.markdown("#### Leads with non-zero payments only")
        st.dataframe(joined_nonzero_export, use_container_width=True)

    with tab_summaries:
        st.markdown("#### Owner Summary")
        st.dataframe(_style_totals_row(_add_totals_row(owner_summary, label_col=owner_col)), use_container_width=True)

        st.markdown("#### Connected Summary")
        st.dataframe(_style_totals_row(_add_totals_row(connected_summary, label_col="Connected")), use_container_width=True)

        st.markdown("#### Self-Converted Detail")
        st.dataframe(self_converted_fact, use_container_width=True)

    with tab_time:
        st.dataframe(_style_totals_row(_add_totals_row(time_summary, label_col="Lead_Created_Date")), use_container_width=True)
        st.pyplot(fig_time, use_container_width=True)

    with tab_hour:
        st.dataframe(_style_totals_row(_add_totals_row(hour_summary, label_col="Lead_Created_Hour")), use_container_width=True)

    with tab_audit:
        st.markdown("### Audit Checks")
        st.dataframe(audit_summary_df, use_container_width=True)

        st.markdown("### Key Counts")
        st.dataframe(audit_counts_df, use_container_width=True)

        st.markdown("### Owner Distribution for Conversion Emails (Deduped First Lead Owner)")
        st.dataframe(owner_dist_df, use_container_width=True)

        st.markdown("### Conversion Email Classification (Full)")
        st.download_button(
            "Download conversion classification CSV",
            data=audit_csv_bytes,
            file_name="conversion_email_classification.csv",
            mime="text/csv",
        )
        st.dataframe(conversion_classification_df, use_container_width=True)

        st.markdown("### Quick Samples")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("Self Converted (first 50)")
            st.dataframe(self_conv_preview_df, use_container_width=True, height=350)
        with c2:
            st.markdown("Sales Conversions (first 50)")
            st.dataframe(sales_conv_preview_df, use_container_width=True, height=350)
        with c3:
            st.markdown("Sales Effort Conversions (first 50)")
            st.dataframe(effort_conv_preview_df, use_container_width=True, height=350)

        st.markdown("### Audit Log Lines")
        if audit_lines:
            for line in audit_lines:
                if "FAIL" in line:
                    st.error(line)
                else:
                    st.info(line)
        else:
            st.write("No audit warnings were generated in this run.")

    with tab_export:
        st.download_button(
            "Download Excel report",
            data=excel_bytes,
            file_name=excel_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with tab_logs:
        if logs:
            for line in logs:
                st.info(line)
        else:
            st.write("No issues logged.")


if __name__ == "__main__":
    main()
